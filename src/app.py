import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from model.Funcionario import Funcionario
import tableformat as tf
from autom import autom
#from dataframes import funcionario as fdf

import model

def finalizar_autom():    
    resposta=""
    while resposta =="":
        resposta = autom.custom_message(f"Deseja finalizar a automação? ", ["sim","não"])
        
    if resposta == "sim": raise autom.AutomException(0,"Programa Interrompido")

def get_problem(text:str="")->str:
    """
    Cria uma janeja perguntando ao usuario se ocorreu algum problema ao tentar execultar alguma operação.
    Se ele escolher "sim", o programa então criará uma janela perguntando a ele qual o problema que aconteceu.

    Returns:
        str: o problema que o usuário relatou, uma string vazia caso nenhum problema aconteceu
    """
    ocorreu_problema = True
    problema= ""
    
    while ocorreu_problema and problema =="":    
        #pergunta se aconteceu algum problema
        ocorreu_problema = autom.custom_message(f"{text}\nOcorreu algum problema? ",["sim", "não"]) == "sim"

        #Caso não tenha ocorrido algum problema
        if not ocorreu_problema:
            break
        
        #caso tenha acontecido algum problema, pedimos que ele informe qual foi o problema que aconteceu
        escolha = autom.custom_message("Informe qual problema aconteceu: ", problems + ["ADICIONAR","LIMPAR"])
    
        if escolha == "ADICIONAR":
            problema = autom.prompt("Adicione um novo erro na lista: ")
            
            if problema =="CANCEL":
                problema=""
                continue
            
            if escolha != "" and escolha not in problems:
                problems.append(problema)
            
        elif escolha == "LIMPAR": #caso ele tenha escolhido limpar os possíveis problemas
            problems.clear()
            continue        
        
        elif escolha =="": #Caso ele tenha fechado a janela, refaça a pergunta
            continue
        
        else:
            problema = escolha     
    
    return problema.upper()


def ask_nedds_alter(func:Funcionario)->bool:
    resposta = ""
    
    while resposta =="":
        resposta = autom.custom_message(f"{func}\n\n É necessário realizar alteração? ", ["sim","não"])
    
        if resposta =="":
            finalizar_autom()
    
    return resposta =="sim"

def confirm_acess(func:Funcionario)-> bool:
    """
    Pergunta ao usuário que ele confirme se foi possível acessar a página de perfil do usuário

    Args:
        func (Funcionario): funcionário no qual será cadastrado

    Returns:
        `bool`: retorna True se o usuário clicou em sim
    """
    resposta= ""
    
    while resposta =="":
        resposta= autom.custom_message(f"Foi possível acessar: {func.nome}?\n ", ["sim","não"])
        
        if resposta =="":
            finalizar_autom()
        
    return resposta =="sim"

#== == == Variables

#col name
sit_DF_col ="SITUACAO_DF"
UNO_col ="UNO"
sit_col ="SITUAÇÃO"

#== Path DataFrame
actual_directory = os.path.dirname(os.path.realpath(__file__))

func_situacao_path = os.path.join(actual_directory,"..","data","func_log.xlsx")

#Dataframe:
func_situacao_df = model.Dataframe(func_situacao_path) 

# == == autom
#moves_to_form
moves_tela_1=[ (560,850), (1312,850), (230,300)]
moves_tela_2=[ (2390,800), (3011,800),(2100,360)]

# 1º Option Input
# 2º Enrollment Input
# 3º Name Input

#options
options=["02"]

moves = moves_tela_1

quanty_press_confirm=3 #quanty_time_press_enter_to_confirm
quanty_press_return=3 #quanty_time_press_f3_to_return

# == == Lista

#== == possibles_problem
problems=[]

#== == situacao
situacao_lista = ["PENDENTE","INALTERADO","ALTERADO"]
#PENDENTE = ainda não foi checado a situação do Funcionario
#ALTERADO = O nome do funcionário não foi alterado  
#ALTERADO = O nome do funcionário foi alterado  

continue_autom= False

#Antes de iniciar...

#Adicioanar a coluna responsavel pela situacao dataframe

#Caso o func_situacao esteja vazio, devemos povoa-lo com os dados da tabela original
if len(func_situacao_df) ==0:
    
    #Consultamos o dataframe original
    func_path = os.path.join(actual_directory,"..","data","func.xlsx")
    func_df = model.Dataframe(func_path)

    for _ , func in func_df.get_iterrows():
        func_situacao_df.add_row(**func)
    
    func_situacao_df.add_column(sit_DF_col,situacao_lista[0])
    func_situacao_df.save()
    
    
autom.alert("Programa Inciado. Por favor, conecte-se a sua conta")

#O usuário deverá escolher em qual tela o SISTEMA 
tela_sis=""
telas_escolha=["1º Monitor","2º Monitor"]

while tela_sis=="":
    tela_sis= autom.custom_message("Em qual monitor está o Sistema?", telas_escolha)

moves = moves_tela_1 if tela_sis == telas_escolha[0] else moves_tela_2
try:

    for index,situacao in func_situacao_df.get_columns(sit_DF_col):
                
        row = func_situacao_df[index]

        matricula = row["MATRICULA"]
        nome_alterar = row["COD_CPF_NOME"] 
        cpf = row["CPF"]
 
        func= Funcionario(matricula,nome_alterar,'',cpf,situacao)
        
        #Caso não esteja PENDENTE a situação do funcionario
        if situacao != situacao_lista[0]:
            print("=="*30)
            print(f"Pulei o funcionário: {func.matricula} : {func.nome}")
            print("=="*30)
            continue        
        #pergunte se deseja continuar a automoção
        continue_autom = False
    
        while not continue_autom:
            continue_autom = autom.confirm(f"Deseja alterar o nome de: {func.nome}?")

            if not continue_autom:
                finalizar_autom()

        #MOVE O CURSOR PARA O INPUT DA OPCAO
        autom.moveCursor(*moves[0])
        autom.click()
        #ESCREVA A OPÇÃO DE ALTERAR
        autom.click()
        autom.write(options[0])
        
        #MOVE O CURSOR PARA O INPUT DA MATRICULA
        autom.moveCursor(*moves[1])
        autom.click()

        #ESCREVA O INPUT DA MATRICULA
        autom.write(matricula)
        
        #CONFIRME
        autom.press_enter()
        
        #Pergunta se foi possível acessar o funcionário
        acess_response = False
        
        while not acess_response:
            #Deverá pedir a confirmação        
            acess_response = confirm_acess(func)
            
            # caso não tenha sido possível acessar
            if not acess_response:

                situacao = get_problem()
                
                if situacao != "" : break #Se tiver tido passado algum problema
                    
        #caso tenha conseguido acessar ao portal do funcionário
        else:
            
            #Pergunte se precisa alterar o funcionário
            needs_alter = ask_nedds_alter(func) 
            if needs_alter:  
                #Continue com a automação
                autom.moveCursor(*moves[2]) #mova o cursor para o input de nome
                autom.click()

                autom.write(nome_alterar)
                autom.press_enter(quanty_press_confirm) #preciona a quantidade de enters necessárias para confirmar
                
                problem = get_problem(f"{func}\n\n Você alterou o nome do funcionário")
                if problem != "":
                    #foi detectado algum erro ao tentar renomear o usuário
                    situacao= problem
                    
                    #peça o codigo do UNO
                    UO_codigo=autom.prompt("Digite o codigo da UNIDADE ORCAMENTARIA (UNO) do funcionário")
                    func_situacao_df.alter_column_index(index,UNO_col,UO_codigo)
                    #peça o codigo da situacao
                    situacao_cod=autom.prompt("Digite o codigo da SITUAÇÃO do funcionário")
                    #altere no dataframe
                    func_situacao_df.alter_column_index(index,sit_col,situacao_cod)
                    autom.press("f3",quanty_press_return)
                else:
                    situacao = situacao_lista[-1]     
                    autom.press_enter(1)
                
            else:
                #CASO NÃO SEJA NECESSÁRIO ALTERAR
                situacao = situacao_lista[1]
        
        #Por fim, alteramos a situação no DF
        func_situacao_df.alter_column_index(index,sit_DF_col,situacao)
       
        print(f"M {matricula} => Nome {nome_alterar}, Sit. {situacao}")

except autom.AutomException as AE:
    print(AE)

except KeyboardInterrupt:
    print("programa interrompido via CMD")

except Exception as E:
    print("Algo incrivelmente incrível aconteceu: ")
    print(E)

func_situacao_df.save()     

# Carregar o arquivo Excel 
workbook = load_workbook(func_situacao_df.path)
ws = workbook.active

sit_df_position = func_situacao_df.columns.get_loc(sit_DF_col)

# == == Collors
#AMARELO quando o funcionário foi devidamente renoemado 
yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

#LARANJA quando não foi necessário renomear
orange = PatternFill(start_color="EC9006", end_color="EC9006", fill_type="solid")

#VERDE quando ocorreu um erro ao tentar renomear o funcionário
red = PatternFill(start_color="FF4200", end_color="FF4200", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row = len(func_situacao_df) + 1, min_col=1, max_col=len(func_situacao_df.columns)):

    situacao_func= row[sit_df_position].value

    if situacao_func == situacao_lista[-1]: #ALTERADO
        tf.paintRow(row,yellow)
        
    elif situacao_func == situacao_lista[1]: #INALTERADO
        tf.paintRow(row,orange)
        
    elif situacao_func == situacao_lista[0]:
        continue
    
    else: #Ocorreu algum problema
        tf.paintRow(row,red)
    
# Salva as alterações no arquivo Excel
workbook.save(func_situacao_df.path)

autom.alert("PROGRAMA FINALIZADO")